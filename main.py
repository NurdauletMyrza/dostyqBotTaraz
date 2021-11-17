import datetime
import openpyxl
import telebot
import emoji


def read_file(directory="home/NurdauletMyrza/telegram dostyq08bot", filename="practice.xlsx"):
    return openpyxl.load_workbook(filename)


def save_file(directory="home/NurdauletMyrza/telegram dostyq08bot/results", filename="practice.xlsx"):
    wb.save(filename)
    print(f"{filename} saved")


def get_keys():
    key_list = list()
    sheet = wb["answers"]
    key_list.append(0)
    while True:
        key = sheet.cell(len(key_list), 1).value
        if not key:
            key_list[0] = len(key_list)
            break
        key_list.append(key)
    return key_list


def get_id_list(sheet):
    data = list()
    while sheet.cell(len(data) + 2, 1).value:
        data.append(sheet.cell(len(data) + 2, 1).value)
    return data


def get_points(row):
    points = dict()
    for i in range(1, keys[0]):
        column = i + 4
        points[i] = wb[group].cell(row, column).value
    return points


def get_new_task(points):
    for task in points.keys():
        if not points[task] and task not in id_task.values():
            return task
    return -1


def set_point(row, column, value):
    if not wb[group].cell(row, column).value:
        n = 0 if value == '+' else 1
        x = int(wb[group].cell(row, 2 + n).value)
        wb[group].cell(row, column, value=value)
        wb[group].cell(row, 2 + n, value=x + 1)
        return True
    return False


def get_result(row):
    right = wb[group].cell(row, column=2).value
    wrong = wb[group].cell(row, column=3).value
    return right, wrong


bot = telebot.TeleBot("1464705930:AAE_eAHD9qndKpMPkkx_5uhgMEBT7ZAuJZ0")
keyboard_get_ticket1 = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
keyboard_get_ticket1.row("Билет алу")
keyboard_get_ticket2 = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
keyboard_get_ticket2.row("Келесі билет")
keyboard_start = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
keyboard_start.row("/start")
keyboard_keys = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
keyboard_keys.row('A', 'B', 'C', 'D', 'E')
keyboard_contact = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
button_phone = telebot.types.KeyboardButton(text="Send my contact", request_contact=True)
keyboard_contact.add(button_phone)


star = emoji.emojize(":star:", use_aliases=True)
hand = emoji.emojize(":hand:", use_aliases=True)
cross_mark = emoji.emojize(":cross_mark:", use_aliases=True)
check_mark = emoji.emojize(":check_mark:", use_aliases=True)


wb = read_file()
keys = get_keys()
group = "results"
id_list = get_id_list(wb[group])
id_task = dict()
id_status = dict()


@bot.message_handler(commands=["start"])
def start_message(message):
    user_id = message.from_user.id
    try:
        bot.send_message(message.chat.id, f"{hand}, {message.from_user.first_name}")
        if user_id in id_list:
            help_message(message)
        else:
            bot.send_message(message.chat.id, "Бастау үшін SEND MY CONTACT батырмасын басыңыз.",
                             reply_markup=keyboard_contact)
    except Exception:
        print(f"start error {user_id}")
        start_message(message)


@bot.message_handler(commands=["help"])
def help_message(message):
    user_id = message.from_user.id
    try:
        bot.send_message(message.chat.id, f"{star}{star}{star}<b>Қош келдіңіз!</b>{star}{star}{star}\n" +
                         '''           Билет алу үшін /ticket
      Нәтижені көру үшін /result
         Ботты тоқатту үшін /end
                         ''', parse_mode='HTML', reply_markup=keyboard_get_ticket1)
    except Exception:
        print(f"help error {user_id}")
        start_message(message)


@bot.message_handler(commands=["ticket"])
def ticket_message(message):
    user_id = message.from_user.id
    try:
        id_status[user_id] = id_status.get(user_id, False)
        if id_status[user_id]:
            id_task[user_id] = get_new_task(get_points(id_list.index(user_id) + 2))
            id_status[user_id] = False
        else:
            id_task[user_id] = id_task.get(user_id, get_new_task(get_points(id_list.index(user_id) + 2)))
        bot.send_message(message.chat.id, f"Билет: {id_task[user_id]}", reply_markup=keyboard_keys)
    except Exception:
        print(f"ticket error {user_id}")
        start_message(message)


@bot.message_handler(commands=["result"])
def result_message(message):
    user_id = message.from_user.id
    try:
        right, wrong = get_result(id_list.index(user_id) + 2)
        bot.send_message(message.chat.id,
                         f"Дұрыс жауап: {right}\nҚате жауап: {wrong}\nКПД: {int(right / (right + wrong) * 100)}%")
    except ZeroDivisionError:
        right, wrong = get_result(id_list.index(user_id) + 2)
        bot.send_message(message.chat.id,
                         f"Дұрыс жауап: {right}\nҚате жауап: {wrong}\nКПД: 0%")
    except Exception:
        print(f"result error {user_id}")
        start_message(message)


@bot.message_handler(commands=["save"])
def save_message(message):
    user_id = message.from_user.id
    try:
        if user_id == 1495184578:
            save_file(filename="practice.xlsx")
            save_file(filename=datetime.datetime.now().strftime("%H-%M") + ".xlsx")
    except Exception:
        print(f"save error {user_id}")
        start_message(message)


@bot.message_handler(commands=["reset"])
def reset_message(message):
    user_id = message.from_user.id
    try:
        if user_id == 1495184578:
            row = 2
            while True:
                if not wb["results"].cell(row, 1).value:
                    print("results deleted")
                    break
                wb["results"].cell(row, 2, value=0)
                wb["results"].cell(row, 3, value=0)
                row = row + 1
    except Exception:
        print(f"reset error {user_id}")
        start_message(message)


@bot.message_handler(commands=["end"])
def end_message(message):
    user_id = message.from_user.id
    try:
        del id_task[user_id]
    except KeyError:
        print(f"KeyError: {user_id}\ndel id_task[user_id]")
    result_message(message)
    bot.send_message(message.chat.id, "Қош болыңыз.", reply_markup=keyboard_start)


@bot.message_handler(content_types=['contact'])
def contact_message(message):
    user_id = message.from_user.id
    try:
        if message.contact is not None:
            print(f"\nname: {message.from_user.first_name}\n"
                  f"id: {user_id}\n"
                  f"phone number: {message.contact.phone_number}\n")
            id_list.append(user_id)
            wb[group].cell(row=len(id_list) + 1, column=1, value=user_id)
            help_message(message)
    except Exception:
        print(f"contact error {user_id}")
        start_message(message)


@bot.message_handler(content_types=["text"])
def text_message(message):
    user_id = message.from_user.id
    try:
        if message.text == "Билет алу" or message.text == "Келесі билет":
            ticket_message(message)
        else:
            answer = message.text
            row = id_list.index(user_id) + 2
            column = id_task[user_id] + 4
            if answer.upper() in "ABCDE" and len(answer) == 1:
                key = keys[id_task[user_id]]
                if answer.lower() == key:
                    if set_point(row, column, '+'):
                        bot.send_message(message.chat.id, f"Дұрыс {check_mark}", reply_markup=keyboard_get_ticket2)
                else:
                    if set_point(row, column, '-'):
                        bot.send_message(message.chat.id, f"Қате {cross_mark}\nДұрыс жауап: {key.upper()}",
                                         reply_markup=keyboard_get_ticket2)
                id_status[user_id] = True
            else:
                bot.send_message(message.chat.id, "Дұрыстап енгізіңіз: A, B, C, D, E")
                ticket_message(message)
    except Exception:
        print(f"text error {user_id}")
        start_message(message)


bot.polling(none_stop=True)
